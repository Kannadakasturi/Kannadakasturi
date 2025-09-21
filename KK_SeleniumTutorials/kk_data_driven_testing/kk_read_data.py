'''
Created on 21-Sept-2025

@author: ADMIN
'''


import openpyxl

#1. Load the workbook
filename = r"C:\Users\ADMIN\Downloads\kasturi.xlsx"
my_workbook = openpyxl.load_workbook(filename)

#2. Get the sheet
active_sheet = my_workbook.active      #to get the active
#active_sheet = my_workbook["Sheet1"]   #to get the sheet by name

#3. Get the no. of rows and columns
total_rows = active_sheet.max_row
print(total_rows)

total_cols = active_sheet.max_column
print(total_cols)

#4. Get the data
'''username = active_sheet.cell(2,1).value
password = active_sheet.cell(2,2).value
print(username, password)'''

username = active_sheet.cell(2,1).value
password = active_sheet.cell(2,2).value
print(username, password)'''


