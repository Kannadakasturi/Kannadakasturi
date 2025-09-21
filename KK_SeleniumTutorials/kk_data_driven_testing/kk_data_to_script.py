'''
Created on 21-Sept-2025

@author: ADMIN
'''
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By

def test_orangehrm_login(username, password, url):
        #1. Launch the Chrome browser with desired capabilities

        options = webdriver.ChromeOptions()
        options.add_experimental_option("detach",True)
        options.add_argument("start-maximized")
        driver = webdriver.Chrome(options=options)
        driver.implicitly_wait(12)

        #2. Navigate to a practice site
        driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")
        
        #3. Enter User name
        user_name_txt_bx = driver.find_element(By.NAME,'username')
        user_name_txt_bx.send_keys(username)

        #4. Enter Password
        password_txt_bx = driver.find_element(By.NAME,"password")
        password_txt_bx.send_keys(password)

        #5. Click on login btn
        login_btn = driver.find_element(By.XPATH,"//button[@type='submit']")
        login_btn.click() 
        
        #6. Validating the login
        expected_url = url
        actual_url = driver.current_url
        if expected_url in actual_url:
            print("test passed")
        else:
            print("test fail")    

#1. Load the workbook
filename = r"C:\Users\ADMIN\Downloads\kasturi.xlsx"
my_workbook = openpyxl.load_workbook(filename)

#2. Get the sheet
active_sheet = my_workbook.active      #to get the active
#active_sheet = my_workbook["Sheet1"]   #to get the sheet by name

#3. Get the no. of rows and columns
total_rows = active_sheet.max_row
print(total_rows)

total_columns = active_sheet.max_column
print(total_columns)

#4. Get the data
for i in range(2, total_rows + 1):  
    username = active_sheet.cell(i, 1).value   
    password = active_sheet.cell(i, 2).value  
    url = active_sheet.cell(i,3).value 
    print(username, password, url) 
    test_orangehrm_login(username, password, url)
    
    

